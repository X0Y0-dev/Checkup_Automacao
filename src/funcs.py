import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import date, timedelta
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule


#region CONFIRGUAÇÃO DE ELEBILIDADE

# Empresas que necessitam de elegibilidade
EMPRESAS = ["Bradesco", "Santander", "Itaú", "Mediservice"]

# Convênios que necessitam de elegibilidade
CONVENIOS = ["Central Nacional Unimed", "Unimed Seguros Saúde", "Care Plus", "Unafisco", "Câmara", "Gama", "Senado", "Intermedici"]

# Opções de status
STATUS_OPC = ["ELEGÍVEL", "PENDENTE", "NÃO ELEGÍVEL", "CANCELADO", "REAGENDADO"]

#endregion

#region CONFIGURAÇÕES DE ESTILIZAÇÃO

# Cabeçalho estilizado na planilha
CINZA_HEADER = PatternFill(
    start_color = "D9D9D9",
    end_color = "D9D9D9",
    fill_type = "solid"
)

# Fonte estilizada do cabeçalho
FONTE_HEADER = Font(bold = True)

# Cor para status ELEGIVEL
verde = PatternFill(
    start_color = "92D050",
    end_color = "92D050",
    fill_type = "solid"
)

# Cor para status PENDENTE
amarelo = PatternFill(
    start_color = "FFD966",
    end_color = "FFD966",
    fill_type = "solid"
)

# Cor para status CANCELADO e NÃO ELEGÍVEL
vermelho = PatternFill(
    start_color = "FF0000",
    end_color = "FF0000",
    fill_type = "solid"
)

# Cor para status REAGENDADO
roxo = PatternFill(
    start_color = "7030A0",
    end_color = "7030A0",
    fill_type = "solid"
)

# Fonte para ELEGÍVEL E PENDENTE
preto = Font(color = "000000")

# Forte para CANCELADO, NÃO ELEGÍVEL E REAGENDADO
branco = Font(color = "FFFFFF")

# Dicionário de estilização de status
STATUS_ESTILOS = {
    "ELEGÍVEL": (verde, preto),
    "PENDENTE": (amarelo, preto),
    "CANCELADO": (vermelho, branco),
    "NÃO ELEGÍVEL": (vermelho, branco),
    "REAGENDADO": (roxo, branco)
}

#endregion

estado = {"df": pd.DataFrame(), "historico": []}

#region FUNÇÕES E TRATAMENTO DE DADOS

def parse_checkup(texto: str) -> pd.DataFrame:

    # Rexes para tratamento de dados dos pacientes
    texto = re.sub(r'[\u200b\u200c\u200d\u202a\u202b\u202c\u202d\u202e\ufeff]', '', texto)
    texto = texto.replace('\r\n', '\n').replace('\r', '\n')
    todas_linhas = [l.strip() for l in texto.split('\n')]

    inicios = [
        i for i, linha in enumerate(todas_linhas)
        if re.match(r'^\d{2}:\d{2}$', linha)
        and i + 1 < len(todas_linhas)
        and not re.match(r'^\d', todas_linhas[i + 1])
    ]

    pacientes = []

    for idx, inicio in enumerate(inicios):

        fim = inicios[idx + 1] if idx + 1 < len(inicios) else len(todas_linhas)
        linhas = [l for l in todas_linhas[inicio:fim] if l]

        if len(linhas) < 6:
            continue

        try:
            hora      = linhas[0]
            paciente  = linhas[1]
            convenio  = linhas[2]
            categoria = linhas[3]

            pacientes.append({
                "Hora":       hora,
                "Paciente":   paciente,
                "Convênio":   convenio,
                "Categoria":  categoria,
                "Status":     "",
            })
        except IndexError:
            pacientes.append({
                "Hora":       linhas[0] if len(linhas) > 0 else "?",
                "Paciente":   linhas[1] if len(linhas) > 1 else "?",
                "Convênio":   linhas[2] if len(linhas) > 4 else "?",
                "Categoria":  linhas[3] if len(linhas) > 6 else "?",
                "Status":     "",
            })
    return pd.DataFrame(pacientes)

def identificar_empresa(convenio: str) -> str | None:
    for empresa in EMPRESAS:
        if empresa.lower() in convenio.lower():
            return empresa
    return None

def identificar_convenio(convenio: str) -> str | None:
    for convs in CONVENIOS:
        if convs.lower() in convenio.lower():
            return convs
    return None

def gerar_txt_convenios(df_convenios: pd.DataFrame) -> bytes:
    linhas = []

    for convenio in sorted(df_convenios["Convênio"].unique()):
        linhas.append(f"\n\n{convenio}:")
        df_conv = df_convenios[df_convenios["Convênio"] == convenio]

        for data in sorted(df_conv["Data"].unique()):
            linhas.append(f"\n  {data}:\n")
            df_data = df_conv[df_conv["Data"] == data]
            for _, row in df_data.iterrows():
                linhas.append(f"    {row['Paciente']} - {row['Categoria']}")
        linhas.append("")

    return "\n".join(linhas).encode("utf-8")  # retorna bytes ao invés de salvar

def gerar_txt_brasilia(df_brasilia: pd.DataFrame) -> bytes:
    linhas = []

    for data in sorted(df_brasilia["Data"].unique()):
        linhas.append(f"{data}:")
        df_data = df_brasilia[df_brasilia["Data"] == data]
        for _, row in df_data.iterrows():
            linhas.append(f"    {row['Paciente']} - {row['Convênio']} - {row['Categoria']}")
        linhas.append("")

    return "\n".join(linhas).encode("utf-8")  # retorna bytes ao invés de salvar

def gerar_txt_empresa(df_empresa: pd.DataFrame, nome_empresa: str) -> bytes:
    linhas = []

    for data in sorted(df_empresa["Data"].unique()):
        linhas.append(f"{data}:")
        df_data = df_empresa[df_empresa["Data"] == data]

        for _, row in df_data.iterrows():
            linhas.append(f"    {row['Paciente']} - {row['Convênio']} - {row['Categoria']}")
        
        linhas.append("")
    
    return "\n".join(linhas).encode("utf-8")

def estilizar_header(worksheet):

    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = CINZA_HEADER
        cell.font = FONTE_HEADER
    
    for column in worksheet.columns:

        max_length = 0
        column_letter = column[0].column_letter
        header = column[0].value

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Largura automárica
        largura = max_length + 2

        # Ajuste de largura para status
        if header == "Status":
            largura = 20
        
        worksheet.column_dimensions[column_letter].width = largura

def estilizar_status(worksheet):
    dv = DataValidation(
        type = "list",
        formula1 = f'"{",".join(STATUS_OPC)}"',
        allow_blank = True
    )

    worksheet.add_data_validation(dv)

    coluna_status = None
    for cell in worksheet[1]:
        if cell.value == "Status":
            coluna_status = cell.column_letter
            break
    
    if not coluna_status:
        return
    
    # Intervalo de validação
    intervalo = f"{coluna_status}2:{coluna_status}{worksheet.max_row}"

    dv.add(intervalo)

    for status, (fill, font) in STATUS_ESTILOS.items():
        worksheet.conditional_formatting.add(
            intervalo,
            FormulaRule(
                formula = [f'{coluna_status}2="{status}"'],
                fill = fill,
                font = font
            )
        )

#endregion